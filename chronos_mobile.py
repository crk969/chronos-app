# File: chronos_mobile.py
# Modulo: Chronos Mobile v2.2 (Definitivo)
# Identificativo: realizzazione CRk969
import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.screenmanager import ScreenManager, Screen, NoTransition
from kivy.clock import Clock
from datetime import datetime, timedelta, date
import json
import os
import calendar
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment

kivy.require('2.0.0')

# --- CLASSE EXCEL GENERATOR (ORA INCLUSA) ---
class ExcelReportGenerator:
    def __init__(self, data_dict):
        self.data = data_dict
        self.workbook = openpyxl.Workbook()
        self.ws = self.workbook.active
        self.ws.title = "Riepilogo Ore"

    def generate_report(self, file_path):
        headers = ["Data", "Tipo Giornata", "Obiettivo Ore", "Ore Permesso", "Ore Lavorate", "Saldo Giornaliero", "Timbrature"]
        self.ws.append(headers)
        
        for date_str, day_data in sorted(self.data.items()):
            timbrature_str = " | ".join([datetime.fromisoformat(ts).strftime('%H:%M') for ts in day_data.get("timbrature", [])])
            worked_seconds = day_data.get("ore_lavorate_sec", 0)
            target_seconds = day_data.get("obiettivo_ore", 0) * 3600
            permit_seconds = day_data.get("ore_permesso", 0) * 3600
            # Saldo = Lavorato + Permesso - Obiettivo
            balance_seconds = worked_seconds + permit_seconds - target_seconds
            
            row = [
                date_str,
                day_data.get("tipo_giornata", ""),
                f"{day_data.get('obiettivo_ore', 0):.2f}",
                f"{day_data.get('ore_permesso', 0):.2f}",
                self._seconds_to_hms(worked_seconds),
                self._seconds_to_hms(balance_seconds, show_sign=True),
                timbrature_str
            ]
            self.ws.append(row)
        
        self._format_sheet()
        self.workbook.save(file_path)

    def _seconds_to_hms(self, seconds, show_sign=False):
        sign = "-" if seconds < 0 else "+" if show_sign else ""
        seconds = int(abs(seconds))
        h, rem = divmod(seconds, 3600); m, s = divmod(rem, 60)
        return f"{sign}{h:02}:{m:02}:{s:02}"
        
    def _format_sheet(self):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in self.ws[1]: cell.font = header_font; cell.fill = header_fill
        for col in self.ws.columns:
            max_length = 0
            for cell in col:
                try: 
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[col[0].column_letter].width = adjusted_width

# --- CLASSE POPUP DI PIANIFICAZIONE ---
class PlannerPopup(Popup):
    DAY_TYPES = ["Lavorativo", "Ferie", "Permesso", "Malattia", "Festività", "Art. 104"]
    ABSENCE_TYPES_HOURLY = ["Permesso", "Art. 104"]
    def __init__(self, app, date_obj, **kwargs):
        super().__init__(**kwargs); self.title = f"Pianifica: {date_obj.strftime('%d/%m/%Y')}"; self.size_hint = (0.95, 0.9)
        self.app = app; self.date_str = date_obj.strftime("%Y-%m-%d"); self.day_data = self.app.data.get(self.date_str, self.app._get_default_day_data(self.date_str))
        main_layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        type_layout = BoxLayout(size_hint_y=None, height='40dp'); type_layout.add_widget(Label(text="Tipo Giornata:"))
        self.spinner_type = Spinner(text=self.day_data.get('tipo_giornata', 'Lavorativo'), values=self.DAY_TYPES); self.spinner_type.bind(text=self.toggle_visibility); type_layout.add_widget(self.spinner_type); main_layout.add_widget(type_layout)
        self.permits_layout = BoxLayout(size_hint_y=None, height='40dp'); self.permits_layout.add_widget(Label(text="Ore Permesso:")); self.hours_input = TextInput(text=str(self.day_data.get('ore_permesso', '0')), multiline=False, input_type='number'); self.permits_layout.add_widget(self.hours_input); main_layout.add_widget(self.permits_layout)
        self.events_layout = BoxLayout(orientation='vertical'); self.events_layout.add_widget(Label(text="Blocchi Orari: definiscono l'obiettivo.", font_size='12sp', size_hint_y=None, height='30dp'))
        self.events_list = GridLayout(cols=1, size_hint_y=None, spacing=5); self.events_list.bind(minimum_height=self.events_list.setter('height')); self.refresh_events_list(); self.events_layout.add_widget(self.events_list)
        add_event_layout = GridLayout(cols=3, size_hint_y=None, height='40dp'); self.start_input = TextInput(hint_text="Inizio (HH:MM)"); self.end_input = TextInput(hint_text="Fine (HH:MM)"); add_btn = Button(text="Aggiungi", on_press=self.add_event)
        add_event_layout.add_widget(self.start_input); add_event_layout.add_widget(self.end_input); add_event_layout.add_widget(add_btn); self.events_layout.add_widget(add_event_layout)
        main_layout.add_widget(self.events_layout)
        save_btn = Button(text="Salva Giorno Singolo", size_hint_y=None, height='48dp', on_press=self.save_changes); main_layout.add_widget(save_btn)
        self.content = main_layout; self.toggle_visibility(None, self.spinner_type.text)
    def toggle_visibility(self, spinner, text):
        if text == "Lavorativo": self.events_layout.height, self.events_layout.opacity = 300, 1; self.permits_layout.height, self.permits_layout.opacity = 0, 0
        elif text in self.ABSENCE_TYPES_HOURLY: self.events_layout.height, self.events_layout.opacity = 0, 0; self.permits_layout.height, self.permits_layout.opacity = '40dp', 1
        else: self.events_layout.height, self.events_layout.opacity = 0, 0; self.permits_layout.height, self.permits_layout.opacity = 0, 0
    def refresh_events_list(self):
        self.events_list.clear_widgets(); events_data = self.day_data.get("eventi_programmati", "[]")
        if isinstance(events_data, str): 
            try: events_data = json.loads(events_data)
            except json.JSONDecodeError: events_data = []
        for start, end in events_data:
            event_row = BoxLayout(size_hint_y=None, height='30dp'); event_row.add_widget(Label(text=f"{start} - {end}"))
            remove_btn = Button(text="X", size_hint_x=None, width='40dp'); remove_btn.bind(on_press=lambda btn, s=start, e=end: self.remove_event(s, e)); event_row.add_widget(remove_btn); self.events_list.add_widget(event_row)
    def add_event(self, instance):
        start, end = self.start_input.text, self.end_input.text
        try:
            datetime.strptime(start, "%H:%M"); datetime.strptime(end, "%H:%M")
            events_data = self.day_data.get("eventi_programmati", "[]")
            if isinstance(events_data, str): events_data = json.loads(events_data)
            events_data.append([start, end]); self.day_data["eventi_programmati"] = json.dumps(sorted(events_data)); self.refresh_events_list(); self.start_input.text = ""; self.end_input.text = ""
        except ValueError: pass
    def remove_event(self, start_time, end_time):
        events_data = self.day_data.get("eventi_programmati", "[]")
        if isinstance(events_data, str): events_data = json.loads(events_data)
        events_data.remove([start_time, end_time]); self.day_data["eventi_programmati"] = json.dumps(events_data); self.refresh_events_list()
    def save_changes(self, instance):
        self.app.update_day_data(self.date_str, self.spinner_type.text, self.day_data.get("eventi_programmati", "[]"), self.hours_input.text)
        self.dismiss(); self.app.planner_screen.calendar_widget.build_calendar()

# --- WIDGET CALENDARIO ---
class CalendarWidget(GridLayout):
    def __init__(self, app, **kwargs):
        super().__init__(**kwargs); self.cols = 7; self.app = app; self.current_date = datetime.now()
        self.build_calendar()
    def build_calendar(self):
        self.clear_widgets(); header = BoxLayout(size_hint_y=None, height='48dp', spacing=10)
        prev_btn = Button(text="<", size_hint_x=0.2, on_press=self.prev_month); self.month_label = Label(text=self.current_date.strftime("%B %Y").upper(), font_size='20sp', bold=True)
        next_btn = Button(text=">", size_hint_x=0.2, on_press=self.next_month); header.add_widget(prev_btn); header.add_widget(self.month_label); header.add_widget(next_btn); self.add_widget(header)
        days = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]
        for day in days: self.add_widget(Label(text=day, bold=True, size_hint_y=None, height='30dp'))
        month_calendar = calendar.monthcalendar(self.current_date.year, self.current_date.month)
        for week in month_calendar:
            for day_date in week:
                if day_date == 0: self.add_widget(Label(text=""))
                else:
                    date_obj = date(self.current_date.year, self.current_date.month, day_date)
                    day_data = self.app.data.get(date_obj.strftime("%Y-%m-%d"), {})
                    day_type = day_data.get('tipo_giornata', None); day_btn = Button(text=str(day_date))
                    if day_type == "Ferie" or day_type == "Festività": day_btn.background_color = (0.2, 0.6, 0.8, 1)
                    elif day_type == "Malattia": day_btn.background_color = (0.8, 0.6, 0.2, 1)
                    elif date_obj == date.today(): day_btn.background_color = (0.5, 0.5, 0.5, 1)
                    day_btn.bind(on_press=self.day_pressed); self.add_widget(day_btn)
    def day_pressed(self, instance):
        day = int(instance.text); date_obj = date(self.current_date.year, self.current_date.month, day); self.app.open_planner_popup(date_obj)
    def prev_month(self, instance): self.current_date = (self.current_date.replace(day=1) - timedelta(days=1)).replace(day=1); self.build_calendar()
    def next_month(self, instance):
        last_day = calendar.monthrange(self.current_date.year, self.current_date.month)[1]; self.current_date = (self.current_date.replace(day=last_day) + timedelta(days=1)); self.build_calendar()

# --- SCHERMATE ---
class ClockScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs); self.app = App.get_running_app(); layout = BoxLayout(orientation='vertical', padding=20, spacing=10)
        self.clock_label = Label(text="--:--:--", font_size='48sp', bold=True); layout.add_widget(self.clock_label)
        self.stamp_button = Button(text="Timbra", size_hint_y=None, height='60dp', font_size='20sp', on_press=self.app.timbra); layout.add_widget(self.stamp_button)
        dashboard_layout = GridLayout(cols=2, size_hint_y=None, height='100dp')
        dashboard_layout.add_widget(Label(text="Lavorato Oggi:", bold=True)); self.worked_today_label = Label(text="00:00:00"); dashboard_layout.add_widget(self.worked_today_label)
        dashboard_layout.add_widget(Label(text="Debito/Credito:", bold=True)); self.balance_label = Label(text="--:--:--"); dashboard_layout.add_widget(self.balance_label); layout.add_widget(dashboard_layout)
        self.stamps_list_label = Label(text="Nessuna timbratura.", size_hint_y=None, height='80dp', halign='center', valign='top'); self.stamps_list_label.bind(size=self.stamps_list_label.setter('text_size')); layout.add_widget(self.stamps_list_label)
        switch_button = Button(text="Vai a Pianificazione >", size_hint_y=None, height='40dp', on_press=self.switch_to_planner); layout.add_widget(switch_button)
        footer = Label(text="realizzazione CRk969 - Dott. Roberto Calò", font_size='10sp', color=(0.7,0.7,0.7,1), size_hint_y=None, height='20dp'); layout.add_widget(footer)
        self.add_widget(layout)
    def switch_to_planner(self, instance): self.manager.current = 'planner'

class PlannerScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs); self.app = App.get_running_app(); layout = BoxLayout(orientation='vertical')
        period_layout = BoxLayout(size_hint_y=None, height='80dp', orientation='vertical', padding=10, spacing=5)
        period_layout.add_widget(Label(text="Applica a un periodo (es. Ferie, Malattia):", bold=True))
        date_selectors = BoxLayout(size_hint_y=None, height='40dp')
        date_selectors.add_widget(Label(text="Dal:")); self.start_date_input = TextInput(hint_text="GG/MM/AAAA", multiline=False); date_selectors.add_widget(self.start_date_input)
        date_selectors.add_widget(Label(text="Al:")); self.end_date_input = TextInput(hint_text="GG/MM/AAAA", multiline=False); date_selectors.add_widget(self.end_date_input)
        apply_btn = Button(text="Applica Assenza a Periodo", on_press=self.apply_period); period_layout.add_widget(date_selectors); period_layout.add_widget(apply_btn); layout.add_widget(period_layout)
        self.calendar_widget = CalendarWidget(app=self.app); layout.add_widget(self.calendar_widget)
        bottom_layout = BoxLayout(size_hint_y=None, height='60dp', padding=10, spacing=10)
        switch_button = Button(text="< Vai a Orologio", on_press=self.switch_to_clock); bottom_layout.add_widget(switch_button)
        help_button = Button(text="Guida (?)", on_press=self.show_help); bottom_layout.add_widget(help_button)
        export_button = Button(text="Esporta Report", on_press=self.app.export_to_excel); bottom_layout.add_widget(export_button)
        layout.add_widget(bottom_layout); self.add_widget(layout)
    def switch_to_clock(self, instance): self.manager.current = 'clock'
    def apply_period(self, instance):
        try:
            start_date = datetime.strptime(self.start_date_input.text, "%d/%m/%Y").date()
            end_date = datetime.strptime(self.end_date_input.text, "%d/%m/%Y").date()
            if end_date < start_date: return
            popup_content = BoxLayout(orientation='vertical', spacing=10); spinner = Spinner(text="Ferie", values=("Ferie", "Malattia")); popup_content.add_widget(spinner); confirm_btn = Button(text="Conferma")
            popup = Popup(title="Seleziona tipo di assenza", content=popup_content, size_hint=(0.6, 0.4))
            def confirm_action(btn_instance):
                day_type = spinner.text; self.app.update_period_data(start_date, end_date, day_type); popup.dismiss(); self.calendar_widget.build_calendar()
            confirm_btn.bind(on_press=confirm_action); popup.open()
        except ValueError: pass
    def show_help(self, instance):
        help_text = ("- Clicca su un giorno per pianificare turni o permessi orari.\n"
                     "- Usa i campi 'Dal'/'Al' per applicare Ferie/Malattia su più giorni.\n"
                     "- L'export genera un file Excel nella cartella dell'app.")
        popup = Popup(title='Guida Pianificazione', content=Label(text=help_text, halign='center'), size_hint=(0.8, 0.4)); popup.open()

# --- APP PRINCIPALE ---
class ChronosMobileApp(App):
    CONFIG_FILE = "chronos_mobile_config.json"; DATA_FILE = "chronos_mobile_data.json"
    def build(self):
        self.config = self._load_json(self.CONFIG_FILE, {"daily_target_hours": 8.5}); self.data = self._load_json(self.DATA_FILE, {})
        self.sm = ScreenManager(transition=NoTransition()); self.clock_screen = ClockScreen(name='clock'); self.planner_screen = PlannerScreen(name='planner')
        self.sm.add_widget(self.clock_screen); self.sm.add_widget(self.planner_screen)
        self.reload_today_data(); Clock.schedule_interval(self.update, 1);
        return self.sm
    def on_stop(self): self._log_day_data()
    def reload_today_data(self):
        self.today_str = datetime.now().strftime("%Y-%m-%d")
        self.today_data = self.data.get(self.today_str, self._get_default_day_data(self.today_str))
        self.timestamps = [datetime.fromisoformat(ts) for ts in self.today_data.get("timbrature", [])]
        self.update_ui_from_state()
    def timbra(self, instance):
        if self.timestamps and (datetime.now() - self.timestamps[-1]).total_seconds() < 1: return
        self.timestamps.append(datetime.now()); self._log_day_data(); self.update_ui_from_state()
    def update(self, dt):
        self.clock_screen.clock_label.text = datetime.now().strftime("%H:%M:%S")
        day_type = self.today_data.get('tipo_giornata', 'Lavorativo')
        if day_type == "Lavorativo":
            worked_seconds = self._calculate_worked_seconds(datetime.now())
            target_seconds = self.today_data.get("obiettivo_ore", 8.5) * 3600
            permit_seconds = self.today_data.get("ore_permesso", 0) * 3600
            remaining_seconds = target_seconds - permit_seconds - worked_seconds
            self.clock_screen.worked_today_label.text = self._seconds_to_hms(worked_seconds)
            self.clock_screen.balance_label.text = self._seconds_to_hms(remaining_seconds)
            self.clock_screen.balance_label.color = (1, 0.2, 0.2, 1) if remaining_seconds > 0 else (0.2, 1, 0.2, 1)
        else:
            self.clock_screen.worked_today_label.text = "00:00:00"; self.clock_screen.balance_label.text = day_type; self.clock_screen.balance_label.color = (0.2, 0.6, 0.8, 1)
    def _is_working(self): return len(self.timestamps) % 2 != 0
    def update_ui_from_state(self):
        if self.today_data.get('tipo_giornata', 'Lavorativo') != 'Lavorativo':
            self.clock_screen.stamp_button.text = self.today_data['tipo_giornata']; self.clock_screen.stamp_button.disabled = True
            self.clock_screen.stamps_list_label.text = "Nessuna timbratura per oggi."; return
        self.clock_screen.stamp_button.disabled = False
        self.clock_screen.stamp_button.text = "Uscita / Pausa" if self._is_working() else "Ingresso / Rientro"
        self.clock_screen.stamp_button.background_color = (0.8, 0.2, 0.2, 1) if self._is_working() else (0.2, 0.8, 0.2, 1)
        stamps_text = [f"{'Ingresso/Rientro' if i % 2 == 0 else 'Uscita/Pausa'}: {ts.strftime('%H:%M:%S')}" for i, ts in enumerate(self.timestamps)]
        self.clock_screen.stamps_list_label.text = "\n".join(stamps_text) if stamps_text else "Nessuna timbratura."
    def _calculate_worked_seconds(self, current_time):
        total_seconds = 0; local_timestamps = self.timestamps[:]
        if self._is_working(): local_timestamps.append(current_time)
        for i in range(0, len(local_timestamps) - 1, 2): total_seconds += (local_timestamps[i+1] - local_timestamps[i]).total_seconds()
        return total_seconds
    def _log_day_data(self):
        worked_seconds = self._calculate_worked_seconds(datetime.now())
        self.today_data['timbrature'] = [ts.isoformat() for ts in self.timestamps]
        self.today_data['ore_lavorate_sec'] = worked_seconds
        self.data[self.today_str] = self.today_data; self._save_data()
    def _load_json(self, file_path, default_data):
        if not os.path.exists(file_path):
            with open(file_path, 'w', encoding='utf-8') as f: json.dump(default_data, f, indent=4)
            return default_data
        try:
            with open(file_path, 'r', encoding='utf-8') as f: return json.load(f)
        except (json.JSONDecodeError, FileNotFoundError): return default_data
    def _save_data(self):
        with open(self.DATA_FILE, 'w', encoding='utf-8') as f: json.dump(dict(sorted(self.data.items())), f, indent=4)
    def _seconds_to_hms(self, seconds, show_sign=False):
        sign = "-" if seconds < 0 else "+" if show_sign else ""; seconds = int(abs(seconds))
        h, rem = divmod(seconds, 3600); m, s = divmod(rem, 60); return f"{sign}{h:02}:{m:02}:{s:02}"
    def calculate_hours_from_events(self, events_list):
        total_minutes = 0
        for start, end in events_list:
            try: t_start = datetime.strptime(start, "%H:%M"); t_end = datetime.strptime(end, "%H:%M"); total_minutes += (t_end - t_start).total_seconds() / 60
            except ValueError: return 0.0
        return round(total_minutes / 60, 2)
    def open_planner_popup(self, date_obj):
        popup = PlannerPopup(app=self, date_obj=date_obj); popup.open()
    def _get_default_day_data(self, date_str):
        return {"tipo_giornata": "Lavorativo", "eventi_programmati": '[["08:30", "13:00"], ["14:00", "18:30"]]', "obiettivo_ore": 8.5, "timbrature": [], "ore_permesso": 0}
    def update_day_data(self, date_str, day_type, events_data, hours_text):
        day_data = self.data.get(date_str, self._get_default_day_data(date_str))
        day_data["tipo_giornata"] = day_type
        if day_type == "Lavorativo":
            if isinstance(events_data, str): events_data = json.loads(events_data)
            day_data["eventi_programmati"] = json.dumps(events_data)
            day_data["obiettivo_ore"] = self.calculate_hours_from_events(events_data); day_data["ore_permesso"] = 0
        elif day_type in PlannerPopup.ABSENCE_TYPES_HOURLY:
            day_data["ore_permesso"] = float(hours_text or 0)
        else:
            day_data["obiettivo_ore"] = 0; day_data["ore_permesso"] = 0
        self.data[date_str] = day_data; self._save_data()
        if date_str == self.today_str: self.reload_today_data()
    def update_period_data(self, start_date, end_date, day_type):
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5:
                date_str = current_date.strftime("%Y-%m-%d")
                day_data = self.data.get(date_str, self._get_default_day_data(date_str))
                day_data['tipo_giornata'] = day_type; day_data['ore_permesso'] = 0; day_data['obiettivo_ore'] = 0
                self.data[date_str] = day_data
            current_date += timedelta(days=1)
        self._save_data()
        if start_date <= date.today() <= end_date: self.reload_today_data()
    def export_to_excel(self, instance):
        try:
            if not self.data:
                popup = Popup(title='Info', content=Label(text='Nessun dato da esportare.'), size_hint=(0.8, 0.4)); popup.open()
                return
            file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"Report_Ore_{datetime.now().strftime('%Y-%m')}.xlsx")
            report = ExcelReportGenerator(self.data)
            report.generate_report(file_path)
            popup = Popup(title='Esportazione Completata', content=Label(text=f"Report salvato in:\n{file_path}", halign='center'), size_hint=(0.8, 0.4)); popup.open()
        except Exception as e:
            popup = Popup(title='Errore Esportazione', content=Label(text=f"Errore: {e}"), size_hint=(0.8, 0.4)); popup.open()

if __name__ == '__main__':
    ChronosMobileApp().run()